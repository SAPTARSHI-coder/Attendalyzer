import json
import re
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# ── Load OCR cache ──────────────────────────────────────────────────────────
with open("ocr_cache.json", "r", encoding="utf-8") as f:
    ocr_cache = json.load(f)

# ── Load Responses sheet (has exact filename submitted per student) ──────────
responses_df = pd.read_excel(
    "B.Tech 4th Semester Attendance Collection (Debarred List) (Responses) (1).xlsx"
)
responses_df.columns = responses_df.columns.str.strip()

# ── Load the master attendance.xlsx (has Sl NO, Name, Roll Number already) ──
att_df = pd.read_excel("attendance.xlsx", header=0)
att_df.columns = att_df.columns.str.strip()

# ── The 12 subjects in order (from attendance.xlsx header row) ──────────────
SUBJECTS_ORDER = [
    "CSE11111",
    "CSE11110",
    "PSG11021",
    "CSE11109",
    "MTH11534",
    "CSE11112",
    "CSE11204",
    "CSE12205",
    "CSE12166",
    "CSE12114",
    "MTH12531",
    "CSE14170",
]

def normalize(s):
    """Lowercase, strip extra spaces/chars for fuzzy matching."""
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def find_cache_entry(name_in_att, roll_in_att):
    """Find the best matching cache entry for a student."""
    n = normalize(name_in_att)

    # 1. Direct match on full name
    for cache_name, data in ocr_cache.items():
        if normalize(cache_name) == n:
            return data, cache_name

    # 2. Partial / word-level match
    words = [w for w in name_in_att.lower().split() if len(w) > 3]
    best_key = None
    best_score = 0
    for cache_name in ocr_cache:
        cn = normalize(cache_name)
        score = sum(1 for w in words if normalize(w) in cn)
        if score > best_score:
            best_score = score
            best_key = cache_name

    if best_key and best_score >= 1:
        return ocr_cache[best_key], best_key

    return None, None


def get_subject_data(subjects_list, code):
    """Return (total, attended) for a given subject code from the extracted list."""
    for s in subjects_list:
        subj_str = s.get("subject", "")
        if code in subj_str.replace(" ", ""):
            tc = s.get("total_classes", 0) or 0
            ac = s.get("attended_classes", 0) or 0
            return tc, ac
    return 0, 0


# ── Build output rows ────────────────────────────────────────────────────────
print("Building attendance matrix...\n")

# We'll build a brand-new DataFrame with the exact columns we need
cols = ["Sl NO", "Name", "Roll Number"]
for code in SUBJECTS_ORDER:
    cols += [f"{code} - T", f"{code} - A", f"{code} - P"]

output_rows = []
missing_students = []

# att_df rows contain each student (header is row 0, data starts from row 1)
name_col = "Name"
roll_col = "Roll Number"
sl_col   = "Sl NO"

for _, row in att_df.iterrows():
    sl   = row[sl_col]
    name = str(row[name_col]).strip()
    roll = str(row[roll_col]).strip()

    if pd.isna(sl) or str(sl).strip() in ("", "nan"):
        continue  # skip empty rows

    cache_data, matched_key = find_cache_entry(name, roll)

    out = {"Sl NO": sl, "Name": name, "Roll Number": roll}

    if cache_data:
        subjects_list = cache_data.get("subject_wise", [])
        print(f"  [{sl}] {name} → matched cache key: '{matched_key}' ({len(subjects_list)} subjects)")
        for code in SUBJECTS_ORDER:
            tc, ac = get_subject_data(subjects_list, code)
            pct = round(ac / tc * 100, 1) if tc > 0 else 0.0
            out[f"{code} - T"] = tc
            out[f"{code} - A"] = ac
            out[f"{code} - P"] = pct
    else:
        print(f"  [{sl}] {name} → ⚠ NO CACHE MATCH")
        missing_students.append(f"{sl}. {name} ({roll})")
        for code in SUBJECTS_ORDER:
            out[f"{code} - T"] = ""
            out[f"{code} - A"] = ""
            out[f"{code} - P"] = ""

    output_rows.append(out)

result_df = pd.DataFrame(output_rows, columns=cols)
print(f"\nTotal rows: {len(result_df)}")
print(f"Missing / unmatched: {len(missing_students)}")
for m in missing_students:
    print("  -", m)


# ── Write to Excel with proper formatting ────────────────────────────────────
OUTPUT_FILE = "Attendance_Matrix_FINAL.xlsx"

# We construct a multi-row header:
#   Row 1: Sl NO | Name | Roll Number | <subject full name merged x3> | ...
#   Row 2: (blank)|(blank)|(blank)    |  T-  |  A-  |  P-  | ...
#   Data starts row 3

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Attendance"

# ── subject full names lookup ────────────────────────────────────────────────
SUBJECT_FULL = {
    "CSE11111": "CSE11111 || Formal Language and Automata",
    "CSE11110": "CSE11110 || Design and Analysis of Algorithms",
    "PSG11021": "PSG11021 || Human Values and Professional Ethics",
    "CSE11109": "CSE11109 || Object Oriented Programming",
    "MTH11534": "MTH11534 || Discrete Structures and Logic",
    "CSE11112": "CSE11112 || Introduction to Artificial Intelligence",
    "CSE11204": "CSE11204 || Exploratory Data Analysis",
    "CSE12205": "CSE12205 || Exploratory Data Analysis Lab",
    "CSE12166": "CSE12166 || Design and Analysis of Algorithms Lab",
    "CSE12114": "CSE12114 || Object Oriented Programming Lab",
    "MTH12531": "MTH12531 || Numerical Techniques Lab",
    "CSE14170": "CSE14170 || Mini Project-I",
}

# Column index helpers
INFO_COLS = 3   # Sl NO, Name, Roll Number
SUBJ_COUNT = len(SUBJECTS_ORDER)
TOTAL_COLS = INFO_COLS + SUBJ_COUNT * 3

# ── Row 1: header ─────────────────────────────────────────────────────────────
hdr_fill   = PatternFill("solid", fgColor="1F4E79")  # dark blue
hdr_font   = Font(bold=True, color="FFFFFF", size=10)
subj_fill  = PatternFill("solid", fgColor="2E75B6")  # medium blue
subj_font  = Font(bold=True, color="FFFFFF", size=9)
t_fill     = PatternFill("solid", fgColor="D6E4F0")
a_fill     = PatternFill("solid", fgColor="FDEBD0")
p_fill     = PatternFill("solid", fgColor="D5F5E3")
center_aln = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_side  = Side(style="thin", color="AAAAAA")
thin_border= Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Info header cells
for ci, label in enumerate(["Sl NO", "Name", "Roll Number"], start=1):
    cell = ws.cell(row=1, column=ci, value=label)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = center_aln
    cell.border = thin_border
    ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

# Subject header cells (merge 3 cols each, then sub-header row)
for si, code in enumerate(SUBJECTS_ORDER):
    start_col = INFO_COLS + si * 3 + 1
    end_col   = start_col + 2

    # Row 1: merged subject name
    cell = ws.cell(row=1, column=start_col, value=SUBJECT_FULL[code])
    cell.fill = subj_fill
    cell.font = subj_font
    cell.alignment = center_aln
    cell.border = thin_border
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    # Row 2: T- / A- / P-
    for offset, sub_label, fill_color in [
        (0, "T-", PatternFill("solid", fgColor="D6DCE4")),
        (1, "A-", PatternFill("solid", fgColor="FCE4D6")),
        (2, "P-", PatternFill("solid", fgColor="E2EFDA")),
    ]:
        c = ws.cell(row=2, column=start_col + offset, value=sub_label)
        c.fill = fill_color
        c.font = Font(bold=True, size=9)
        c.alignment = center_aln
        c.border = thin_border

ws.row_dimensions[1].height = 42
ws.row_dimensions[2].height = 16

# ── Data rows ─────────────────────────────────────────────────────────────────
green_fill  = PatternFill("solid", fgColor="C6EFCE")
red_fill    = PatternFill("solid", fgColor="FFC7CE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")

for ri, r in enumerate(output_rows, start=3):
    ws.cell(row=ri, column=1, value=r["Sl NO"]).alignment = center_aln
    ws.cell(row=ri, column=2, value=r["Name"])
    ws.cell(row=ri, column=3, value=r["Roll Number"]).alignment = center_aln

    # Alternating row shade
    row_bg = PatternFill("solid", fgColor="F2F2F2" if ri % 2 == 0 else "FFFFFF")
    for ci in range(1, INFO_COLS + 1):
        ws.cell(row=ri, column=ci).fill = row_bg
        ws.cell(row=ri, column=ci).border = thin_border

    for si, code in enumerate(SUBJECTS_ORDER):
        start_col = INFO_COLS + si * 3 + 1
        tc  = r.get(f"{code} - T", "")
        ac  = r.get(f"{code} - A", "")
        pct = r.get(f"{code} - P", "")

        c_t = ws.cell(row=ri, column=start_col,     value=tc  if tc  != "" else "N/A")
        c_a = ws.cell(row=ri, column=start_col + 1, value=ac  if ac  != "" else "N/A")
        c_p = ws.cell(row=ri, column=start_col + 2, value=f"{pct}%" if pct != "" else "N/A")

        for c in (c_t, c_a, c_p):
            c.alignment = center_aln
            c.border = thin_border

        # Colour the percentage cell
        if isinstance(pct, float):
            if pct >= 75:
                c_p.fill = green_fill
                c_p.font = Font(color="276221", bold=True, size=9)
            elif pct >= 65:
                c_p.fill = yellow_fill
                c_p.font = Font(color="9C5700", bold=True, size=9)
            else:
                c_p.fill = red_fill
                c_p.font = Font(color="9C0006", bold=True, size=9)

    ws.row_dimensions[ri].height = 15

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 20
for si in range(SUBJ_COUNT):
    start_col = INFO_COLS + si * 3 + 1
    ws.column_dimensions[get_column_letter(start_col)].width     = 6
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 6
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 7

# Freeze top 2 rows + first 3 columns
ws.freeze_panes = "D3"

wb.save(OUTPUT_FILE)
print(f"\n✅  Saved: {OUTPUT_FILE}")

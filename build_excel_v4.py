"""
Builds the final Excel from ocr_cache_v2.json
Includes exact course code matching and Medical Certificate data.
"""
import json, re, sys, os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

# ── The 12 subjects in the ORDER matching attendance.xlsx cols 7-18 ────────────
SUBJECTS = [
    ("CSE11111", "Formal Language and Automata"),
    ("CSE11110", "Design and Analysis of Algorithms"),
    ("PSG11021", "Human Values and Professional Ethics"),
    ("CSE11109", "Object Oriented Programming"),
    ("MTH11534", "Discrete Structures and Logic"),
    ("CSE11112", "Introduction to Artificial Intelligence"),
    ("CSE11204", "Exploratory Data Analysis"),
    ("CSE12205", "Exploratory Data Analysis Lab"),
    ("CSE12166", "Design and Analysis of Algorithms Lab"),
    ("CSE12114", "Object Oriented Programming Lab"),
    ("MTH12531", "Numerical Techniques Lab"),
    ("CSE14170", "Mini Project-I"),
]

# ── Load data ─────────────────────────────────────────────────────────────────
with open("ocr_cache_v2.json", encoding='utf-8') as f:
    cache = json.load(f)

att_df = pd.read_excel("attendance.xlsx", header=0)
att_df.columns = att_df.columns.str.strip()

med_df = pd.read_excel("B.Tech 4th Semester Attendance Collection (Debarred List) (Responses) (1).xlsx")
med_map = {}
for _, row in med_df.iterrows():
    r = str(row.get("Student's University Roll Number", "")).strip().lower()
    med_prov = str(row.get("Medical certificate provided?", "")).strip()
    med_range = str(row.get("Medical certificate range written in certificate", "")).strip()
    if r and med_prov and med_prov.lower() != "nan":
        med_map[r] = {"prov": med_prov, "range": med_range if med_range.lower() != "nan" else ""}

NAME_COL = "Name"
ROLL_COL = "Roll Number"
SL_COL   = "Sl NO"

def normalize_code(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def find_subject(subjects_list, target_code):
    tc_norm = normalize_code(target_code)
    for s in subjects_list:
        sc = normalize_code(s.get("code", ""))
        if sc == tc_norm:
            tc  = int(s.get("total", 0) or 0)
            ac  = int(s.get("present", 0) or 0)
            ac  = min(ac, tc)
            pct = round(ac / tc * 100, 1) if tc > 0 else 0.0
            return tc, ac, pct
    return 0, 0, 0.0

def find_cache(name):
    if name in cache: return cache[name]
    nl = name.lower()
    for k in cache:
        if k.lower() == nl: return cache[k]
    words = [w for w in name.lower().split() if len(w) > 3]
    best_k, best_s = None, 0
    for k in cache:
        kl = k.lower()
        score = sum(1 for w in words if w in kl)
        if score > best_s:
            best_s, best_k = score, k
    if best_k and best_s >= 1:
        return cache[best_k]
    return []

# ── Build row data ────────────────────────────────────────────────────────────
rows = []
for _, row in att_df.iterrows():
    sl   = row[SL_COL]
    name = str(row[NAME_COL]).strip()
    roll = str(row[ROLL_COL]).strip()
    if pd.isna(sl) or str(sl).strip() in ("", "nan"):
        continue

    subjects_list = find_cache(name)
    n_found = len(subjects_list)
    print(f"  [{int(sl):02d}] {name:<30} → {n_found} subjects", flush=True)

    roll_norm = roll.lower()
    med_info = med_map.get(roll_norm, {"prov": "No", "range": ""})

    rec = {"sl": int(sl), "name": name, "roll": roll, "med_prov": med_info["prov"], "med_range": med_info["range"]}
    for code, _ in SUBJECTS:
        tc, ac, pct = find_subject(subjects_list, code)
        rec[code] = (tc, ac, pct)
    rows.append(rec)

# ── Create Excel ──────────────────────────────────────────────────────────────
OUT = "Attendance_Matrix_FINAL_v4.xlsx"
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Subject-wise Attendance"

# --- styles ---
def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=9): return Font(bold=bold, color=hex_, size=sz)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left",   vertical="center", wrap_text=False)
thin = Side(style="thin", color="BBBBBB")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FILL  = fill("1F4E79"); HDR_FONT  = font("FFFFFF", bold=True, sz=10)
SUBJ_FILL = fill("2E75B6"); SUBJ_FONT = font("FFFFFF", bold=True, sz=8)
T_FILL    = fill("DCE6F1"); A_FILL    = fill("FCE4D6"); P_FILL = fill("E2EFDA")
GRN_FILL  = fill("C6EFCE"); YLW_FILL  = fill("FFEB9C"); RED_FILL = fill("FFC7CE")
GRN_FONT  = font("276221", bold=True); YLW_FONT = font("9C5700", bold=True); RED_FONT = font("9C0006", bold=True)
ALT_FILL  = fill("F7FAFF")

INFO_COLS  = 3
N_SUBJ     = len(SUBJECTS)
TOTAL_COLS = INFO_COLS + N_SUBJ * 3

# ── Row 1: headers ─────────────────────────────────────────────────────────────
for ci, lbl in enumerate(["Sl NO", "Name", "Roll Number"], 1):
    c = ws.cell(row=1, column=ci, value=lbl)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = ctr; c.border = bdr
    ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

for si, (code, name_) in enumerate(SUBJECTS):
    sc = INFO_COLS + si * 3 + 1
    c  = ws.cell(row=1, column=sc, value=f"{code} || {name_}")
    c.fill = SUBJ_FILL; c.font = SUBJ_FONT; c.alignment = ctr; c.border = bdr
    ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc+2)

    for off, lbl, fl in [(0,"T-",T_FILL), (1,"A-",A_FILL), (2,"P-",P_FILL)]:
        c2 = ws.cell(row=2, column=sc+off, value=lbl)
        c2.fill = fl; c2.font = Font(bold=True, size=9); c2.alignment = ctr; c2.border = bdr

MED_COL_START = INFO_COLS + N_SUBJ * 3 + 1
c = ws.cell(row=1, column=MED_COL_START, value="Medical Certificate Provided?")
c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = ctr; c.border = bdr
ws.merge_cells(start_row=1, start_column=MED_COL_START, end_row=2, end_column=MED_COL_START)

c = ws.cell(row=1, column=MED_COL_START+1, value="Medical Date Range")
c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = ctr; c.border = bdr
ws.merge_cells(start_row=1, start_column=MED_COL_START+1, end_row=2, end_column=MED_COL_START+1)

ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 15

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri, rec in enumerate(rows, start=3):
    is_alt = ri % 2 == 0
    row_fill = ALT_FILL if is_alt else fill("FFFFFF")

    c = ws.cell(row=ri, column=1, value=rec["sl"])
    c.alignment = ctr; c.border = bdr; c.fill = row_fill

    c = ws.cell(row=ri, column=2, value=rec["name"])
    c.alignment = lft; c.border = bdr; c.fill = row_fill

    c = ws.cell(row=ri, column=3, value=rec["roll"])
    c.alignment = ctr; c.border = bdr; c.fill = row_fill

    for si, (code, _) in enumerate(SUBJECTS):
        sc = INFO_COLS + si * 3 + 1
        tc, ac, pct = rec[code]

        ct = ws.cell(row=ri, column=sc,   value=tc if tc else "")
        ca = ws.cell(row=ri, column=sc+1, value=ac if ac else "")
        cp = ws.cell(row=ri, column=sc+2, value=f"{pct}%" if tc else "N/A")

        for cx in (ct, ca, cp):
            cx.alignment = ctr; cx.border = bdr; cx.fill = row_fill

        if tc:
            if pct >= 75:
                cp.fill = GRN_FILL; cp.font = GRN_FONT
            elif pct >= 65:
                cp.fill = YLW_FILL; cp.font = YLW_FONT
            else:
                cp.fill = RED_FILL; cp.font = RED_FONT

    cm1 = ws.cell(row=ri, column=MED_COL_START, value=rec.get("med_prov", ""))
    cm2 = ws.cell(row=ri, column=MED_COL_START+1, value=rec.get("med_range", ""))
    for cx in (cm1, cm2):
        cx.alignment = ctr; cx.border = bdr; cx.fill = row_fill

    ws.row_dimensions[ri].height = 14

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 20
for si in range(N_SUBJ):
    sc = INFO_COLS + si * 3 + 1
    ws.column_dimensions[get_column_letter(sc)].width     = 5
    ws.column_dimensions[get_column_letter(sc+1)].width   = 5
    ws.column_dimensions[get_column_letter(sc+2)].width   = 6

ws.column_dimensions[get_column_letter(MED_COL_START)].width = 20
ws.column_dimensions[get_column_letter(MED_COL_START+1)].width = 30

ws.freeze_panes = "D3"

wb.save(OUT)
print(f"\n✅  Saved: {OUT}  ({len(rows)} students)", flush=True)

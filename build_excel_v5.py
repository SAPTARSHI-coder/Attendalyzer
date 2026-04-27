"""
Builds the final Excel from ocr_cache_v2.json
Includes:
 - 12-subject attendance matrix with T/A/P columns
 - Medical Certificate data
 - Missed Attendance claims (raw text) — with rule: only 1 subject per date gets granted
 - Corrected Attendance column per subject (OCR + granted corrections)
 - Corrected Overall % column
 - Event Participation details (data only, no attendance granted)
"""
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

SUBJECTS = [
    ("CSE11111", "FLAT"),
    ("CSE11110", "DAA"),
    ("PSG11021", "Ethics"),
    ("CSE11109", "OOP"),
    ("MTH11534", "Discrete"),
    ("CSE11112", "AI"),
    ("CSE11204", "EDA"),
    ("CSE12205", "EDA Lab"),
    ("CSE12166", "DAA Lab"),
    ("CSE12114", "OOP Lab"),
    ("MTH12531", "NT Lab"),
    ("CSE14170", "Mini Project"),
]
CODE_MAP = {code: sname for code, sname in SUBJECTS}
SNAME_LIST = [s for _, s in SUBJECTS]

# ── Load OCR cache ─────────────────────────────────────────────────────────────
with open("ocr_cache_v2.json", encoding="utf-8") as f:
    cache = json.load(f)

# ── Load attendance.xlsx (master list) ────────────────────────────────────────
att_df = pd.read_excel("attendance.xlsx", header=0)
att_df.columns = att_df.columns.str.strip()
NAME_COL = "Name"
ROLL_COL = "Roll Number"
SL_COL   = "Sl NO"

# ── Load form responses ────────────────────────────────────────────────────────
form_df = pd.read_excel("B.Tech 4th Semester Attendance Collection (Debarred List) (Responses) (1).xlsx")
COL_ROLL   = "Student's University Roll Number"
COL_MISS   = "Any attendance missed but you are present?"
COL_MED_P  = "Medical certificate provided?"
COL_MED_R  = "Medical certificate range written in certificate"
COL_EVENT  = "Events participation details with date"

# Build maps keyed by roll (lower)
med_map   = {}
miss_map  = {}
event_map = {}

for _, row in form_df.iterrows():
    r = str(row.get(COL_ROLL, "")).strip().lower()
    if not r or r == "nan":
        continue

    med_p = str(row.get(COL_MED_P, "")).strip()
    med_r = str(row.get(COL_MED_R, "")).strip()
    med_map[r] = {
        "prov":  med_p  if med_p  not in ("nan","")  else "No",
        "range": med_r  if med_r  not in ("nan","")  else ""
    }

    miss = str(row.get(COL_MISS, "")).strip()
    junk = {"yes","no","nah","n","nothing","yes ","everything is correct."}
    miss_map[r] = miss if miss.lower() not in junk else ""

    ev = str(row.get(COL_EVENT, "")).strip()
    event_map[r] = ev if ev.lower() not in ("nan","no","nah","nothing","") else ""

# ── Helpers ────────────────────────────────────────────────────────────────────
def normalize_code(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def find_subject_data(subjects_list, target_code):
    tc_norm = normalize_code(target_code)
    for s in subjects_list:
        sc = normalize_code(s.get("code",""))
        if sc == tc_norm:
            tc  = int(s.get("total",  0) or 0)
            ac  = int(s.get("present",0) or 0)
            ac  = min(ac, tc)
            pct = round(ac/tc*100, 1) if tc > 0 else 0.0
            return tc, ac, pct
    return 0, 0, 0.0

def find_cache(name):
    if name in cache: return cache[name]
    nl = name.lower()
    for k in cache:
        if k.lower() == nl: return cache[k]
    words = [w for w in nl.split() if len(w) > 3]
    best_k, best_s = None, 0
    for k in cache:
        kl = k.lower()
        score = sum(1 for w in words if w in kl)
        if score > best_s:
            best_s, best_k = score, k
    if best_k and best_s >= 1: return cache[best_k]
    return []

# ── Parse missed-attendance text → {date: [subject_shortname, ...]} ───────────
SUBJ_KEYWORDS = {
    "CSE11111": ["flat","formal language","formal lang","automata","fla"],
    "CSE11110": ["daa","design and analysis of algo","design & analysis","algorithms"],
    "PSG11021": ["ethics","human values","hvpe","professional ethics"],
    "CSE11109": ["oop","object oriented programming","oops","java","object-oriented"],
    "MTH11534": ["discrete","discrete math","discrete structure","dsl","discrete structures"],
    "CSE11112": ["ai","intro to ai","introduction to artificial","artificial intelligence"],
    "CSE11204": ["eda","exploratory data analysis","data analysis"],
    "CSE12205": ["eda lab","exploratory data analysis lab","data analysis lab"],
    "CSE12166": ["daa lab","algorithms lab","design and analysis of algorithms lab"],
    "CSE12114": ["oop lab","oops lab","object oriented programming lab","java lab"],
    "MTH12531": ["nt lab","numerical techniques lab","numerical lab","numerical techniques"],
    "CSE14170": ["mini project","project"],
}

def match_subject_from_text(text):
    """Return list of (code, shortname) that appear in the text snippet."""
    tl = text.lower()
    found = []
    for code, kws in SUBJ_KEYWORDS.items():
        for kw in kws:
            if kw in tl:
                found.append(code)
                break
    return list(set(found))

# Very simple date extractor: finds dd/mm or dd-mm-yy patterns and English month names
import re as _re

MONTH_MAP = {
    "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
    "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12
}

def extract_date_subject_pairs(text):
    """
    Returns dict: { 'DD/MM' : [list of codes on that date] }
    Rule applied: if a date has > 1 subject, it is rejected (not granted).
    """
    if not text:
        return {}

    # Split by common delimiters: newline, semicolon, &&, 'and'
    # Each segment may describe one date + one or more subjects
    segments = _re.split(r"[;\n]|&&", text)
    date_subjects = defaultdict(list)

    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue

        # Try to extract a date from the segment
        dates_found = []

        # Pattern: dd/mm/yy or dd/mm or dd.mm.yy or dd.mm
        for m in _re.finditer(r"(\d{1,2})[/.\-](\d{1,2})(?:[/.\-]\d{2,4})?", seg):
            d, mo = int(m.group(1)), int(m.group(2))
            if 1 <= d <= 31 and 1 <= mo <= 12:
                dates_found.append(f"{d:02d}/{mo:02d}")

        # Pattern: "9th April", "27th March", "16th January"
        for m in _re.finditer(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)", seg):
            d = int(m.group(1))
            mon_str = m.group(2)[:3].lower()
            mo = MONTH_MAP.get(mon_str)
            if mo and 1 <= d <= 31:
                dates_found.append(f"{d:02d}/{mo:02d}")

        if not dates_found:
            continue

        # Extract subjects from the same segment
        codes = match_subject_from_text(seg)

        for dt in dates_found:
            date_subjects[dt].extend(codes)

    # Deduplicate
    for dt in date_subjects:
        date_subjects[dt] = list(set(date_subjects[dt]))

    return dict(date_subjects)

def apply_grant_rule(date_subjects):
    """
    Returns list of (date, code) that are GRANTED (only 1 subject on that date).
    Dates with 2+ subjects on same day are rejected.
    """
    granted = []
    for dt, codes in date_subjects.items():
        if len(codes) == 1:
            granted.append((dt, codes[0]))
        # else: 2+ subjects same date → rejected
    return granted

# ── Manual overrides (admin corrections beyond auto-parsing) ───────────────────
# Format: { "Name as in attendance.xlsx" : { "SUBJECT_CODE": extra_classes_to_add } }
MANUAL_OVERRIDES = {
    "Shubhajit Mandal": {
        "CSE12114": 6,   # OOP Lab — manually granted by admin
    },
}

# ── Build rows ─────────────────────────────────────────────────────────────────
rows = []
for _, row in att_df.iterrows():
    sl   = row[SL_COL]
    name = str(row[NAME_COL]).strip()
    roll = str(row[ROLL_COL]).strip()
    if pd.isna(sl) or str(sl).strip() in ("","nan"):
        continue

    subjects_list = find_cache(name)
    roll_norm = roll.lower()

    med  = med_map.get(roll_norm,   {"prov":"No","range":""})
    miss = miss_map.get(roll_norm,  "")
    ev   = event_map.get(roll_norm, "")

    # Parse auto-granted corrections
    date_subjects = extract_date_subject_pairs(miss)
    granted = apply_grant_rule(date_subjects)           # list of (date, code)
    granted_codes = defaultdict(int)
    for _, code in granted:
        granted_codes[code] += 1                        # +1 per granted date per subject

    # Apply manual overrides on top of auto-grants
    for override_name, overrides in MANUAL_OVERRIDES.items():
        if name.lower() == override_name.lower():
            for code, extra in overrides.items():
                granted_codes[code] += extra

    rec = {
        "sl": int(sl), "name": name, "roll": roll,
        "med_prov": med["prov"], "med_range": med["range"],
        "missed_raw": miss, "event": ev,
        "granted": dict(granted_codes),
    }

    total_classes_all  = 0
    total_attended_all = 0

    for code, sname in SUBJECTS:
        tc, ac, pct = find_subject_data(subjects_list, code)
        grant = granted_codes.get(code, 0)
        corrected_ac  = min(ac + grant, tc)            # never exceed total
        corrected_pct = round(corrected_ac/tc*100, 1) if tc > 0 else 0.0
        rec[code] = (tc, ac, pct, corrected_ac, corrected_pct, grant)
        total_classes_all  += tc
        total_attended_all += corrected_ac

    rec["overall_pct"] = round(total_attended_all/total_classes_all*100, 1) if total_classes_all > 0 else 0.0
    print(f"  [{int(sl):02d}] {name:<30} granted={dict(granted_codes)}", flush=True)
    rows.append(rec)

# ── Create Excel ──────────────────────────────────────────────────────────────
OUT = "Attendance_Matrix_FINAL_v6.xlsx"
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Subject-wise Attendance"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=9): return Font(bold=bold, color=hex_, size=sz)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="BBBBBB")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FILL  = fill("1F4E79"); HDR_FONT  = font("FFFFFF", bold=True, sz=10)
SUBJ_FILL = fill("2E75B6"); SUBJ_FONT = font("FFFFFF", bold=True, sz=8)
CORR_FILL = fill("375623"); CORR_FONT = font("FFFFFF", bold=True, sz=8)
T_FILL    = fill("DCE6F1"); A_FILL    = fill("FCE4D6"); P_FILL = fill("E2EFDA")
CP_FILL   = fill("D6E4BC")
GRN_FILL  = fill("C6EFCE"); YLW_FILL  = fill("FFEB9C"); RED_FILL = fill("FFC7CE")
GRN_FONT  = font("276221", bold=True); YLW_FONT = font("9C5700", bold=True); RED_FONT = font("9C0006", bold=True)
ALT_FILL  = fill("F7FAFF")
MED_FILL  = fill("FFF2CC")
MISS_FILL = fill("FCE4D6")
EV_FILL   = fill("DDEBF7")

INFO_COLS  = 3
N_SUBJ     = len(SUBJECTS)

# ── Row 1 & 2 headers ─────────────────────────────────────────────────────────
def hdr_cell(r, c, val, merge_r=None, merge_c=None, bg=HDR_FILL, ft=HDR_FONT):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = bg; cell.font = ft; cell.alignment = ctr; cell.border = bdr
    if merge_r or merge_c:
        ws.merge_cells(
            start_row=r, start_column=c,
            end_row=merge_r or r, end_column=merge_c or c
        )
    return cell

# Fixed left cols
for ci, lbl in enumerate(["Sl NO","Name","Roll Number"], 1):
    hdr_cell(1, ci, lbl, merge_r=2)

col = INFO_COLS
# Subject blocks: T-, A-, P-, Corrected A-, Corrected P- = 5 cols each
for si, (code, sname) in enumerate(SUBJECTS):
    sc = INFO_COLS + si*5 + 1
    hdr_cell(1, sc, f"{code} | {sname}", merge_c=sc+4, bg=SUBJ_FILL, ft=SUBJ_FONT)
    for off, lbl, fl in [(0,"T-",T_FILL),(1,"A-",A_FILL),(2,"P-",P_FILL),(3,"Corr A",CP_FILL),(4,"Corr P",CP_FILL)]:
        c2 = ws.cell(row=2, column=sc+off, value=lbl)
        c2.fill = fl; c2.font = Font(bold=True, size=8, color="000000"); c2.alignment = ctr; c2.border = bdr

TAIL_START = INFO_COLS + N_SUBJ * 5 + 1
hdr_cell(1, TAIL_START,   "Overall Corr %",   merge_r=2, bg=CORR_FILL, ft=CORR_FONT)
hdr_cell(1, TAIL_START+1, "Medical?",         merge_r=2, bg=MED_FILL,  ft=font("000000",bold=True))
hdr_cell(1, TAIL_START+2, "Medical Dates",    merge_r=2, bg=MED_FILL,  ft=font("000000",bold=True))
hdr_cell(1, TAIL_START+3, "Attendance Claimed Missed\n(Raw — 1 subject/date rule applied)", merge_r=2, bg=MISS_FILL, ft=font("000000",bold=True))
hdr_cell(1, TAIL_START+4, "Event Participation Details", merge_r=2, bg=EV_FILL,  ft=font("000000",bold=True))

ws.row_dimensions[1].height = 45
ws.row_dimensions[2].height = 18

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri, rec in enumerate(rows, start=3):
    row_fill = ALT_FILL if ri % 2 == 0 else fill("FFFFFF")

    ws.cell(row=ri, column=1, value=rec["sl"]).alignment   = ctr
    ws.cell(row=ri, column=2, value=rec["name"]).alignment = lft
    ws.cell(row=ri, column=3, value=rec["roll"]).alignment = ctr
    for ci in range(1, 4):
        ws.cell(ri, ci).border = bdr; ws.cell(ri, ci).fill = row_fill

    for si, (code, sname) in enumerate(SUBJECTS):
        sc = INFO_COLS + si*5 + 1
        tc, ac, pct, corr_ac, corr_pct, grant = rec[code]

        vals = [tc or "", ac or "", f"{pct}%" if tc else "N/A",
                corr_ac or "", f"{corr_pct}%" if tc else "N/A"]
        for off, v in enumerate(vals):
            cx = ws.cell(row=ri, column=sc+off, value=v)
            cx.alignment = ctr; cx.border = bdr; cx.fill = row_fill

        # colour code original P-
        cp = ws.cell(ri, sc+2)
        if tc:
            if pct >= 75: cp.fill = GRN_FILL; cp.font = GRN_FONT
            elif pct >= 65: cp.fill = YLW_FILL; cp.font = YLW_FONT
            else: cp.fill = RED_FILL; cp.font = RED_FONT

        # colour code corrected P-
        cc = ws.cell(ri, sc+4)
        if tc:
            if corr_pct >= 75: cc.fill = GRN_FILL; cc.font = GRN_FONT
            elif corr_pct >= 65: cc.fill = YLW_FILL; cc.font = YLW_FONT
            else: cc.fill = RED_FILL; cc.font = RED_FONT

    # Tail cells
    ov = ws.cell(ri, TAIL_START, value=f"{rec['overall_pct']}%")
    ov.alignment = ctr; ov.border = bdr
    if rec["overall_pct"] >= 75: ov.fill = GRN_FILL; ov.font = GRN_FONT
    elif rec["overall_pct"] >= 65: ov.fill = YLW_FILL; ov.font = YLW_FONT
    else: ov.fill = RED_FILL; ov.font = RED_FONT

    for off, val, fl in [
        (1, rec["med_prov"],   MED_FILL),
        (2, rec["med_range"],  MED_FILL),
        (3, rec["missed_raw"], MISS_FILL),
        (4, rec["event"],      EV_FILL),
    ]:
        cx = ws.cell(ri, TAIL_START+off, value=val)
        cx.fill = fl; cx.alignment = lft; cx.border = bdr

    ws.row_dimensions[ri].height = 30

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 18
for si in range(N_SUBJ):
    sc = INFO_COLS + si*5 + 1
    for off, w in enumerate([5, 5, 6, 6, 6]):
        ws.column_dimensions[get_column_letter(sc+off)].width = w

for off, w in enumerate([10, 14, 28, 50, 40]):
    ws.column_dimensions[get_column_letter(TAIL_START+off)].width = w

ws.freeze_panes = "D3"
wb.save(OUT)
print(f"\n✅  Saved: {OUT}  ({len(rows)} students)", flush=True)
